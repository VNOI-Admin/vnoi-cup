#!/usr/bin/env python3

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
import pandas as pd

QUALIFY_FROM_R1 = 3
QUALIFY_FROM_R2 = 4

QUALIFIED = [
    "huyhau6a2",
    "BaoJiaoPisu",
    "duongnc000",
    "daoquanglinh2007",
    "huutuan",
    "cpbeginner",
    "DeMen100ns",
    "Pannda",
    "khanhphucscratch",
    "SanguineChameleon",
    "buidangnguyen05",
    "neko_nyaaaaaaaaaaaaaaaaaaaaaaa",
    "hollwo_pelw",
    "vinaka",
    "flashmt",
    "thieunguyenhuy",
    "Mediocrity",
    "toidaidot2",
    "hxano",
    "powder",
    "nathan4690",
    "unknownnaciul",
    "LeKienThanh69",
    "minhcool",
]

EXCLUDED = [
    "maomao90",
    "Zanite",
    "DylanSmith",
    "asiad",
    "PhawinBot",
    "toxicpie",
    "ajinkya1p3",
    "IceKnight1093",
    "Ra16bit",
    "kenti",
    "errorgorn",
    "physics0523",
]


def get_sec(time_str):
    h, m, s = time_str.split(":")
    return int(h) * 3600 + int(m) * 60 + int(s)


def main():
    finalist = []

    def check_qualify(df, to_qualify):
        rank = 1
        for idx, row in df.iterrows():
            if row["Username"] in finalist:
                df.at[idx, "Rank"] = ""
            else:
                df.at[idx, "Rank"] = rank
                if rank <= to_qualify:
                    finalist.append(row["Username"])
                rank += 1

    r1 = pd.read_csv("ranking_r1.csv")
    check_qualify(r1, QUALIFY_FROM_R1)
    r1["Penalty"] = r1["Penalty"].fillna("00:00:00").apply(get_sec)
    r1 = r1[[c for c in r1 if c != "Rank"] + ["Rank"]]
    r1 = r1.rename(
        columns={"Points": "R1 Points", "Penalty": "R1 Penalty", "Rank": "R1 Rank"}
    )

    r2 = pd.read_csv("ranking_r2.csv")
    check_qualify(r2, QUALIFY_FROM_R2)
    r2["Penalty"] = r2["Penalty"].fillna("00:00:00").apply(get_sec)
    r2 = r2[[c for c in r2 if c != "Rank"] + ["Rank"]]
    r2 = r2.rename(
        columns={"Points": "R2 Points", "Penalty": "R2 Penalty", "Rank": "R2 Rank"}
    )

    total = r1.merge(r2, on=["Username", "Full Name"], how="outer")
    total["R1 Points"] = total["R1 Points"].fillna(0)
    total["R2 Points"] = total["R2 Points"].fillna(0)
    total["R1 Penalty"] = total["R1 Penalty"].fillna(0)
    total["R2 Penalty"] = total["R2 Penalty"].fillna(0)
    total = total[(total["R1 Points"] >= 0) & (total["R2 Points"] >= 0)]
    total["Total Points"] = total["R1 Points"] + total["R2 Points"]
    total["Total Penalty"] = total["R1 Penalty"] + total["R2 Penalty"]
    total = total.sort_values(
        by=["Total Points", "Total Penalty"], ascending=[False, True]
    )
    total.reset_index(drop=True, inplace=True)
    total.index.name = "Rank"
    total.index += 1

    total["Qualified"] = ""
    for idx, row in total.iterrows():
        if row["Username"] in QUALIFIED:
            total.at[idx, "Qualified"] = "x"

    column_names = [
        "Username",
        "Full Name",
        "Total Points",
        "Total Penalty",
        "Qualified",
    ]
    column_names = column_names + [c for c in total if c not in column_names]
    total = total[column_names]
    total["Full Name"] = total["Full Name"].apply(
        lambda x: "'" + str(x) if str(x).startswith("=") else x
    )
    total.to_excel("total.xlsx")

    keep_columns = ["Username", "Total Points"]
    tshirt_candidates = total[(total["Qualified"] == "") & (total["Total Points"] > 0)]
    tshirt_candidates = tshirt_candidates.drop(
        columns=[c for c in tshirt_candidates if c not in keep_columns]
    )
    tshirt_candidates = tshirt_candidates[~tshirt_candidates["Username"].isin(EXCLUDED)]
    tshirt_candidates.to_csv("tshirt_candidates.csv")

    wb = load_workbook(filename="total.xlsx")
    ws = wb[wb.sheetnames[0]]

    for row in ws.rows:
        if row[5].value == "x":
            print(row[1].value)
            for cell in row:
                cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)) + 1)
                )

    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.save("total.xlsx")


if __name__ == "__main__":
    main()
